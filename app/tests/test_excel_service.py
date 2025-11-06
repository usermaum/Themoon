"""
ExcelSyncService 테스트

Excel 동기화 및 마이그레이션 기능 테스트
"""

import pytest
import os
from datetime import date, timedelta
from app.services.excel_service import ExcelSyncService
from app.services.roasting_service import RoastingService
from app.models.database import Bean, RoastingLog


@pytest.fixture
def sample_roasting_data(db_session):
    """
    테스트용 로스팅 데이터 생성
    """
    # 원두 생성
    bean = Bean(
        no=1,
        name="테스트원두",
        country_code="KR",
        roast_level="medium",
        price_per_kg=5000,
        status="active"
    )
    db_session.add(bean)
    db_session.commit()

    # 2025-10월 로스팅 기록 3개 생성
    test_date = date(2025, 10, 15)
    logs = []

    for i in range(3):
        log = RoastingService.create_roasting_log(
            db=db_session,
            raw_weight_kg=100.0 + i * 10,
            roasted_weight_kg=83.0 + i * 8,
            roasting_date=test_date - timedelta(days=i),
            expected_loss_rate=17.0
        )
        logs.append(log)

    return {
        'bean': bean,
        'logs': logs,
        'month': '2025-10'
    }


class TestExcelExport:
    """Excel 내보내기 테스트"""

    def test_export_roasting_logs_success(self, db_session, sample_roasting_data, tmp_path):
        """
        로스팅 기록 Excel 내보내기 성공
        """
        # Given
        month = sample_roasting_data['month']
        output_file = str(tmp_path / f"{month}_테스트.xlsx")

        # When
        result_path = ExcelSyncService.export_roasting_logs_to_excel(
            db=db_session,
            month=month,
            output_path=output_file
        )

        # Then
        assert result_path == output_file
        assert os.path.exists(result_path)
        assert os.path.getsize(result_path) > 0

        # 파일 정리
        os.remove(result_path)

    def test_export_empty_month(self, db_session, tmp_path):
        """
        데이터가 없는 월 내보내기
        """
        # Given
        month = '2099-12'  # 데이터가 없는 월
        output_file = str(tmp_path / f"{month}_empty.xlsx")

        # When
        result_path = ExcelSyncService.export_roasting_logs_to_excel(
            db=db_session,
            month=month,
            output_path=output_file
        )

        # Then - 데이터가 없으면 None 반환
        assert result_path is None

    def test_export_invalid_month_format(self, db_session, tmp_path):
        """
        잘못된 월 형식
        """
        # Given
        invalid_month = '202510'  # YYYY-MM 형식이 아님
        output_file = str(tmp_path / "invalid.xlsx")

        # When/Then
        # 잘못된 형식이지만 쿼리는 실행됨 (결과 없음 → None 반환)
        result_path = ExcelSyncService.export_roasting_logs_to_excel(
            db=db_session,
            month=invalid_month,
            output_path=output_file
        )

        # 데이터가 없으므로 None 반환
        assert result_path is None


class TestExcelDataValidation:
    """Excel 데이터 검증 테스트"""

    def test_validate_exported_data(self, db_session, sample_roasting_data, tmp_path):
        """
        내보낸 Excel 데이터 검증
        """
        # Given
        month = sample_roasting_data['month']
        output_file = str(tmp_path / f"{month}_validation.xlsx")

        # When
        result_path = ExcelSyncService.export_roasting_logs_to_excel(
            db=db_session,
            month=month,
            output_path=output_file
        )

        # Then - 파일 존재 및 크기 확인
        assert os.path.exists(result_path)
        assert os.path.getsize(result_path) > 1000  # 최소 1KB 이상

        # Excel 파일 읽기 및 검증
        try:
            from openpyxl import load_workbook

            wb = load_workbook(result_path)
            ws = wb.active

            # 헤더 확인 (최소 1행)
            assert ws.max_row >= 1

            # 데이터 행 확인 (헤더 + 3개 데이터)
            assert ws.max_row >= 4  # 헤더 + 3개 로그

            wb.close()
        except ImportError:
            pytest.skip("openpyxl이 설치되지 않음")

        # 파일 정리
        os.remove(result_path)

    def test_exported_file_structure(self, db_session, sample_roasting_data, tmp_path):
        """
        내보낸 파일 구조 검증
        """
        # Given
        month = sample_roasting_data['month']
        output_file = str(tmp_path / f"{month}_structure.xlsx")

        # When
        result_path = ExcelSyncService.export_roasting_logs_to_excel(
            db=db_session,
            month=month,
            output_path=output_file
        )

        # Then
        try:
            from openpyxl import load_workbook

            wb = load_workbook(result_path)
            ws = wb.active

            # 컬럼 개수 확인 (최소 5개: 날짜, 생두량, 로스팅량, 손실률, 비고)
            assert ws.max_column >= 5

            # 워크시트 이름 확인
            assert ws.title

            wb.close()
        except ImportError:
            pytest.skip("openpyxl이 설치되지 않음")

        # 파일 정리
        os.remove(result_path)


class TestExcelEdgeCases:
    """Excel 서비스 경계 조건 테스트"""

    def test_export_large_dataset(self, db_session, tmp_path):
        """
        대량 데이터 내보내기 (100개)
        """
        # Given - 100개 로스팅 기록 생성
        test_date = date(2025, 10, 15)

        for i in range(100):
            RoastingService.create_roasting_log(
                db=db_session,
                raw_weight_kg=100.0,
                roasted_weight_kg=83.0,
                roasting_date=test_date,
                expected_loss_rate=17.0
            )

        # When
        output_file = str(tmp_path / "large_dataset.xlsx")
        result_path = ExcelSyncService.export_roasting_logs_to_excel(
            db=db_session,
            month='2025-10',
            output_path=output_file
        )

        # Then
        assert os.path.exists(result_path)
        assert os.path.getsize(result_path) > 5000  # 최소 5KB 이상

        try:
            from openpyxl import load_workbook

            wb = load_workbook(result_path)
            ws = wb.active

            # 100개 데이터 + 헤더
            assert ws.max_row >= 101

            wb.close()
        except ImportError:
            pytest.skip("openpyxl이 설치되지 않음")

        # 파일 정리
        os.remove(result_path)

    def test_export_special_characters(self, db_session, tmp_path):
        """
        특수 문자가 포함된 데이터 내보내기
        """
        # Given
        bean = Bean(
            no=99,
            name="특수!@#$%원두",
            country_code="한국",
            roast_level="dark",
            price_per_kg=5000,
            status="active"
        )
        db_session.add(bean)
        db_session.commit()

        log = RoastingService.create_roasting_log(
            db=db_session,
            raw_weight_kg=100.0,
            roasted_weight_kg=83.0,
            roasting_date=date(2025, 10, 15),
            expected_loss_rate=17.0
        )

        # When
        output_file = str(tmp_path / "special_chars.xlsx")
        result_path = ExcelSyncService.export_roasting_logs_to_excel(
            db=db_session,
            month='2025-10',
            output_path=output_file
        )

        # Then
        assert os.path.exists(result_path)

        # 파일 정리
        os.remove(result_path)

    def test_export_with_default_path(self, db_session, sample_roasting_data):
        """
        기본 경로로 내보내기 (output_path=None)
        """
        # Given
        month = sample_roasting_data['month']

        # When
        result_path = ExcelSyncService.export_roasting_logs_to_excel(
            db=db_session,
            month=month,
            output_path=None  # 기본 경로 사용
        )

        # Then
        assert result_path is not None

        # 기본 경로인지 확인 (Data/ 디렉토리)
        if os.path.exists(result_path):
            assert 'Data/' in result_path or month in result_path
            # 파일 정리
            os.remove(result_path)


class TestExcelValidation:
    """Excel 검증 기능 테스트"""

    def test_validate_phase1_migration_success(self, db_session, sample_roasting_data):
        """
        Phase 1 마이그레이션 검증 성공
        """
        # Given - sample_roasting_data에 이미 3개 로그 존재

        # When
        result = ExcelSyncService.validate_phase1_migration(db_session)

        # Then
        assert result['total_logs'] == 3
        assert result['validation_passed'] is True
        assert result['checks']['raw_weight_valid'] == 3
        assert result['checks']['roasted_weight_valid'] == 3
        assert result['checks']['loss_rate_valid'] == 3
        assert result['checks']['no_null_dates'] == 3
        assert len(result['errors']) == 0

    def test_validate_phase1_migration_empty(self, db_session):
        """
        마이그레이션 검증 - 데이터 없음
        """
        # Given - 빈 데이터베이스

        # When
        result = ExcelSyncService.validate_phase1_migration(db_session)

        # Then
        assert result['total_logs'] == 0
        assert result['checks']['raw_weight_valid'] == 0

    def test_validate_phase1_migration_invalid_data(self, db_session):
        """
        마이그레이션 검증 - 잘못된 데이터 감지
        """
        # Given - 잘못된 손실률 (음수)
        invalid_log = RoastingLog(
            raw_weight_kg=100.0,
            roasted_weight_kg=120.0,  # 로스팅량 > 생두량 (비정상)
            roasting_date=date(2025, 10, 15),
            roasting_month='2025-10',
            loss_rate_percent=-20.0,  # 음수 (비정상)
            expected_loss_rate_percent=17.0,
            loss_variance_percent=0.0
        )
        db_session.add(invalid_log)
        db_session.commit()

        # When
        result = ExcelSyncService.validate_phase1_migration(db_session)

        # Then
        assert result['total_logs'] == 1
        assert result['validation_passed'] is False
        assert len(result['errors']) > 0  # 오류 발견

    def test_get_migration_summary_with_data(self, db_session, sample_roasting_data):
        """
        마이그레이션 요약 - 데이터 있음
        """
        # Given - sample_roasting_data에 3개 로그 존재

        # When
        summary = ExcelSyncService.get_migration_summary(db_session)

        # Then
        assert summary['roasting_logs'] == 3
        assert summary['beans'] >= 1  # 최소 1개 원두
        assert summary['total_raw_weight_kg'] > 0
        assert summary['total_roasted_weight_kg'] > 0
        assert summary['avg_loss_rate_percent'] > 0
        assert '✓ 마이그레이션 완료' in summary['status']

    def test_get_migration_summary_empty(self, db_session):
        """
        마이그레이션 요약 - 데이터 없음
        """
        # Given - 빈 데이터베이스

        # When
        summary = ExcelSyncService.get_migration_summary(db_session)

        # Then
        assert summary['roasting_logs'] == 0
        assert summary['beans'] == 0
        assert summary['total_raw_weight_kg'] == 0
        assert summary['total_roasted_weight_kg'] == 0
        assert summary['avg_loss_rate_percent'] == 0
        assert '⏳ 데이터 없음' in summary['status']

    def test_validate_phase1_migration_duplicate_dates(self, db_session):
        """
        마이그레이션 검증 - 중복 날짜 감지
        """
        # Given - 같은 날짜에 2개의 로그 생성
        test_date = date(2025, 10, 15)

        for i in range(2):
            log = RoastingLog(
                raw_weight_kg=100.0 + i * 10,
                roasted_weight_kg=83.0,
                roasting_date=test_date,  # 동일한 날짜
                roasting_month='2025-10',
                loss_rate_percent=17.0,
                expected_loss_rate_percent=17.0,
                loss_variance_percent=0.0
            )
            db_session.add(log)

        db_session.commit()

        # When
        result = ExcelSyncService.validate_phase1_migration(db_session)

        # Then
        assert result['total_logs'] == 2
        assert len(result['errors']) > 0  # 중복 날짜 에러 발견

        # 중복 날짜 메시지 확인
        error_messages = ' '.join(result['errors'])
        assert '중복 날짜' in error_messages or '2025-10-15' in error_messages


class TestExcelIntegration:
    """Excel 서비스 통합 테스트"""

    def test_export_and_verify_workflow(self, db_session, sample_roasting_data, tmp_path):
        """
        전체 워크플로우: 데이터 생성 → 내보내기 → 검증
        """
        # Given
        month = sample_roasting_data['month']
        logs_count = len(sample_roasting_data['logs'])

        # When - 내보내기
        output_file = str(tmp_path / f"{month}_workflow.xlsx")
        result_path = ExcelSyncService.export_roasting_logs_to_excel(
            db=db_session,
            month=month,
            output_path=output_file
        )

        # Then - 검증
        assert os.path.exists(result_path)

        try:
            from openpyxl import load_workbook

            wb = load_workbook(result_path)
            ws = wb.active

            # 데이터 개수 확인
            assert ws.max_row >= logs_count + 1  # 헤더 + 데이터

            # 첫 번째 데이터 행 확인 (헤더 다음)
            if ws.max_row > 1:
                first_data_row = 2
                assert ws.cell(row=first_data_row, column=1).value is not None

            wb.close()
        except ImportError:
            pytest.skip("openpyxl이 설치되지 않음")

        # 파일 정리
        os.remove(result_path)
