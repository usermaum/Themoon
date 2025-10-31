"""
ExcelSyncService: Excel 동기화 및 마이그레이션 서비스

로스팅 데이터를 Excel로 내보내고 마이그레이션을 검증합니다.
"""

from sqlalchemy.orm import Session
from sqlalchemy import func
from app.models.database import RoastingLog, Bean, Blend, BlendRecipe
from datetime import datetime
import logging
import os

logger = logging.getLogger(__name__)


class ExcelSyncService:
    """Excel 동기화 및 마이그레이션 서비스"""

    @staticmethod
    def export_roasting_logs_to_excel(
        db: Session,
        month: str,
        output_path: str = None
    ) -> str:
        """
        월별 로스팅 기록을 Excel로 내보내기

        Args:
            db: SQLAlchemy 세션
            month: 조회 월 (YYYY-MM 형식)
            output_path: 저장 경로 (기본값: Data/{month}_로스팅.xlsx)

        Returns:
            저장된 파일 경로
        """

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            logger.error("❌ openpyxl이 설치되지 않았습니다")
            raise ImportError("openpyxl이 필요합니다. pip install openpyxl을 실행하세요")

        # 로스팅 기록 조회
        logs = db.query(RoastingLog).filter(
            RoastingLog.roasting_month == month
        ).order_by(RoastingLog.roasting_date).all()

        if not logs:
            logger.warning(f"⚠️ {month}의 로스팅 데이터가 없습니다")
            return None

        # 파일 경로 설정
        if not output_path:
            os.makedirs("Data", exist_ok=True)
            output_path = f"Data/{month}_로스팅.xlsx"

        # Workbook 생성
        wb = Workbook()
        ws = wb.active
        ws.title = f"{month}_로스팅"

        # 헤더 설정
        headers = ['날짜', '생두투입(kg)', '로스팅량(kg)', '손실률(%)', '예상손실률(%)', '편차(%)', '비고']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF", size=12)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 데이터 입력
        for row, log in enumerate(logs, 2):
            ws.cell(row=row, column=1, value=log.roasting_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=round(log.raw_weight_kg, 1))
            ws.cell(row=row, column=3, value=round(log.roasted_weight_kg, 1))
            ws.cell(row=row, column=4, value=round(log.loss_rate_percent, 2))
            ws.cell(row=row, column=5, value=round(log.expected_loss_rate_percent, 2))
            ws.cell(row=row, column=6, value=round(log.loss_variance_percent, 2) if log.loss_variance_percent else 0)
            ws.cell(row=row, column=7, value=log.notes or '')

        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 25

        # 파일 저장
        wb.save(output_path)
        logger.info(f"✓ 로스팅 기록 내보내기: {output_path} ({len(logs)}건)")

        return output_path

    @staticmethod
    def validate_phase1_migration(db: Session) -> dict:
        """
        Phase 1 마이그레이션 검증

        Args:
            db: SQLAlchemy 세션

        Returns:
            검증 결과 딕셔너리
        """

        # RoastingLog 데이터 조회
        logs = db.query(RoastingLog).all()

        validations = {
            'total_logs': len(logs),
            'checks': {
                'raw_weight_valid': 0,
                'roasted_weight_valid': 0,
                'loss_rate_valid': 0,
                'no_null_dates': 0,
                'no_duplicates': 0
            },
            'errors': []
        }

        if not logs:
            logger.warning("⚠️ 검증할 로스팅 기록이 없습니다")
            return validations

        # 1. 무게 유효성 확인
        for log in logs:
            # 생두 투입량과 로스팅량이 모두 양수인지 확인
            if log.raw_weight_kg > 0 and log.roasted_weight_kg > 0:
                validations['checks']['raw_weight_valid'] += 1

            # 로스팅량 <= 생두 투입량 확인 (손실 발생)
            if log.roasted_weight_kg <= log.raw_weight_kg:
                validations['checks']['roasted_weight_valid'] += 1
            else:
                validations['errors'].append(
                    f"로그 {log.id}: 로스팅량({log.roasted_weight_kg}kg) > "
                    f"생두투입량({log.raw_weight_kg}kg)"
                )

            # 2. 손실률 검증 (0~50% 범위)
            if 0 <= log.loss_rate_percent <= 50:
                validations['checks']['loss_rate_valid'] += 1
            else:
                validations['errors'].append(
                    f"로그 {log.id}: 손실률 이상 ({log.loss_rate_percent}%)"
                )

            # 3. 날짜 검증
            if log.roasting_date:
                validations['checks']['no_null_dates'] += 1

        # 4. 중복 검증 (같은 날짜의 여러 기록)
        duplicates = db.query(
            RoastingLog.roasting_date,
            func.count().label('count')
        ).group_by(RoastingLog.roasting_date).having(
            func.count() > 1
        ).all()

        if duplicates:
            validations['errors'].append(f"중복 날짜 {len(duplicates)}개 발견")
            for dup in duplicates:
                validations['errors'].append(f"  • {dup[0]}: {dup[1]}건")
        else:
            validations['checks']['no_duplicates'] = len(logs)

        # 최종 검증 결과
        validations['validation_passed'] = len(validations['errors']) == 0

        # 로깅
        if validations['validation_passed']:
            logger.info(f"✓ Phase 1 마이그레이션 검증 통과: {len(logs)}건 모두 유효")
        else:
            logger.warning(f"⚠️ Phase 1 마이그레이션 검증 실패: {len(validations['errors'])}개 오류")

        return validations

    @staticmethod
    def get_migration_summary(db: Session) -> dict:
        """
        마이그레이션 요약 정보 반환

        Args:
            db: SQLAlchemy 세션

        Returns:
            마이그레이션 요약 정보
        """

        logs = db.query(RoastingLog).all()
        beans = db.query(Bean).all()
        blends = db.query(Blend).all()
        recipes = db.query(BlendRecipe).all()

        if logs:
            total_raw = sum(log.raw_weight_kg for log in logs)
            total_roasted = sum(log.roasted_weight_kg for log in logs)
            avg_loss = (total_raw - total_roasted) / total_raw * 100 if total_raw > 0 else 0
        else:
            total_raw = total_roasted = avg_loss = 0

        return {
            'timestamp': datetime.now().isoformat(),
            'roasting_logs': len(logs),
            'beans': len(beans),
            'blends': len(blends),
            'recipes': len(recipes),
            'total_raw_weight_kg': round(total_raw, 2),
            'total_roasted_weight_kg': round(total_roasted, 2),
            'avg_loss_rate_percent': round(avg_loss, 2),
            'status': '✓ 마이그레이션 완료' if logs else '⏳ 데이터 없음'
        }
